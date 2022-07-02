import datetime
from typing import List, Tuple, Optional

from openpyxl import load_workbook

from employee import Employee, CPFEntry, EmploymentStatus


class InvalidSheetError(Exception):
    pass


def load_work_sheet(filename, *args, **kwargs):
    wb = load_workbook(filename, data_only=True)
    if kwargs.get('month', None):
        current_month = kwargs.get('month')
    else:
        current_month = datetime.datetime.now().strftime('%b')
    try:
        ws = wb[current_month]
        wb.close()
        return ws
    except KeyError as e:
        raise InvalidSheetError from e


class SheetParserSM:
    Pending, Begin, End = range(3)


def get_employees_from_sheet(ws):
    employees: List[Tuple] = []
    state = SheetParserSM.Pending
    for row in ws.iter_rows(min_row=1, max_col=30):
        second_col = row[1].value.upper() if row[1].value else ''

        if state == SheetParserSM.Pending:
            if 'NAME' in second_col:
                state = SheetParserSM.Begin
        elif state == SheetParserSM.Begin:
            if 'GRAND TOTAL' in second_col:
                state = SheetParserSM.End
            elif second_col:
                employees.append(row)
        else:
            break
    return employees


def create_cpf_entry(row):
    return CPFEntry(
        ordinary_wage=row[7].value,
        addtional_wage=row[10].value,
        agency_fund=abs(row[17].value)
    )


def parse_dt(input_dt: datetime.datetime) -> str:
    if input_dt.year < 1980:
        modified_dt = datetime.datetime(year=input_dt.year + 100, month=input_dt.month, day=input_dt.day)
        return modified_dt.strftime('%d.%b.%Y')
    else:
        return input_dt.strftime('%d.%b.%Y')


def parse_birth_dt(birth_dt_str, birth_yr) -> str:
    birth_dt = datetime.datetime.strptime(birth_dt_str, '%d/%m')
    return f'{birth_dt.strftime("%d.%b")}.{birth_yr}'


def parse_date_left(input_dt: datetime.datetime) -> Optional[str]:
    if input_dt:
        return parse_dt(input_dt)
    return None


def get_employment_status(input_dt: datetime.datetime) -> EmploymentStatus:
    if input_dt:
        return EmploymentStatus.Left
    return EmploymentStatus.Existing


def create_employee(row):
    start_dt = parse_dt(row[20].value)
    birth_dt = parse_birth_dt(row[23].value, row[24].value)
    date_left = parse_date_left(row[21].value)
    employment_status = get_employment_status(row[21].value)
    e = Employee(
        id_number='',
        name=row[1].value,
        citizenship='',
        start_date=start_dt,
        employment_status=employment_status,
        date_left=date_left,
        date_of_birth=birth_dt,
        cpf_contribution_type='',
        sdl_payable=True
    )
    e.set_cpf_entry(create_cpf_entry(row))
    return e


if __name__ == '__main__':
    ws = load_work_sheet('./TnT-Salary-2022.xlsx')
    rows = get_employees_from_sheet(ws)
    employees = [create_employee(row) for row in rows]
