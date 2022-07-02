from typing import List, Tuple

from openpyxl import load_workbook
from datetime import datetime

from csv_generator import Output
from employee_cpf_record import EmployeeCPFRecord, Agency


class InputFileConfig:
    def __init__(self, sheet_name=None, column_index_mapping=None, desired_rows_config=None,
                 additional_handlers={}) -> None:
        self.sheet_name = sheet_name or datetime.now().strftime('%b')
        self.column_index_mapping = column_index_mapping
        self.desired_rows_config = desired_rows_config or (1, 'NAME', 'TOTAL')
        self.additional_handlers = additional_handlers


class InvalidSheetError(Exception):
    pass


class SheetParserSM:
    Pending, Begin, End = range(3)


class InputFile:
    def __init__(self, filename, config: InputFileConfig) -> None:
        self.filename = filename
        self.config = config

    def load_work_sheet(self):
        wb = load_workbook(self.filename, data_only=True)
        try:
            ws = wb[self.config.sheet_name]
            wb.close()
            return ws
        except KeyError as e:
            raise InvalidSheetError from e

    def get_employees_from_sheet(self):
        ws = self.load_work_sheet()
        employees: List[Tuple] = []
        state = SheetParserSM.Pending
        target_col_index, desired_rows_start_identifier, desired_rows_end_identifier = self.config.desired_rows_config
        for row in ws.iter_rows(min_row=1, max_col=30):
            target_col = row[target_col_index].value.upper() if row[target_col_index].value else ''

            if state == SheetParserSM.Pending:
                if desired_rows_start_identifier in target_col:
                    state = SheetParserSM.Begin
            elif state == SheetParserSM.Begin:
                if desired_rows_end_identifier in target_col:
                    state = SheetParserSM.End
                elif target_col:
                    employees.append(row)
            else:
                break
        return employees

    def create_record(self, row) -> EmployeeCPFRecord:
        return EmployeeCPFRecord(
            name=self.process_row_value(row, Output.name),
            id_number=self.process_row_value(row, Output.id_number),
            ordinary_wage=self.process_row_value(row, Output.ordinary_wage),
            additional_wage=self.process_row_value(row, Output.additional_wage),
            agency_fund=self.process_row_value(row, Output.agency_fund),
            agency=self.process_row_value(row, Output.agency) or Agency.CDAC,
            citizenship=self.process_row_value(row, Output.citizenship),
            pr_start_date=self.process_row_value(row, Output.pr_start_date),
            cpf_contribution_type=self.process_row_value(row, Output.cpf_contribution_type),
            employment_status=self.process_row_value(row, Output.employment_status),
            date_left=self.process_row_value(row, Output.date_left),
            date_of_birth=self.process_row_value(row, Output.date_of_birth),
            sdl_payable=self.process_row_value(row, Output.sdl_payable) or True,
        )

    def process_row_value(self, row, column_name):
        raws = self.extract_row_value(row, column_name)

        # handle for >1 raw value
        if len(raws) > 1:
            if column_name in self.config.additional_handlers:
                return self.config.additional_handlers.get(column_name)(*raws)
            else:
                raise NotImplementedError()

        # default: 1 raw value
        raw = raws[0]
        if column_name in self.config.additional_handlers:
            return self.config.additional_handlers.get(column_name)(raw)
        return raw

    def extract_row_value(self, row, column_name) -> List:
        try:
            if column_name in self.config.column_index_mapping:
                index = self.config.column_index_mapping.get(column_name)
                if isinstance(index, list):
                    return [row[i].value for i in index]
                elif isinstance(index, int):
                    return [row[index].value]
                else:
                    raise IndexError()
        except Exception as e:
            raise e
        else:
            return [None]

    def process(self) -> List[EmployeeCPFRecord]:
        employees = self.get_employees_from_sheet()
        records = [self.create_record(employee) for employee in employees]
        return records
